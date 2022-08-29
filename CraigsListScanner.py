# CraigsListScanner.py

# With Notepad++, use F5 then copy this into box
# C:\Users\Casey\Anaconda3\python.exe -i "$(FULL_CURRENT_PATH)"

import os
path = os.path.abspath(os.path.dirname(__file__))
from bs4 import BeautifulSoup
import re
import requests 
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import os
import string
import datetime
# from forex_python.converter import CurrencyRates
import emoji
import re

# ------------------------------------------------------
# DEFINE FUNCTIONS
# ------------------------------------------------------

def fontStyle(cell):
	fontObj1 = Font(name='Times New Roman', size=10)
	cell.font = fontObj1
	cell.border = border
	cell.alignment = Alignment(horizontal='center', vertical='center')
		
def fontStyleURL(cell, link):
	fontObj1 = Font(name='Times New Roman', size=10)
	cell.value = 'View link!'
	cell.style = "Hyperlink"
	cell.font = fontObj1
	cell.border = border
	cell.hyperlink = link
		
# def remove_emojis(data):
    # emoj = re.compile("["
        # u"\U0001F600-\U0001F64F"  # emoticons
        # u"\U0001F300-\U0001F5FF"  # symbols & pictographs
        # u"\U0001F680-\U0001F6FF"  # transport & map symbols
        # u"\U0001F1E0-\U0001F1FF"  # flags (iOS)
        # u"\U00002500-\U00002BEF"  # chinese char
        # u"\U00002702-\U000027B0"
        # u"\U00002702-\U000027B0"
        # u"\U000024C2-\U0001F251"
        # u"\U0001f926-\U0001f937"
        # u"\U00010000-\U0010ffff"
        # u"\u2640-\u2642" 
        # u"\u2600-\u2B55"
        # u"\u200d"
        # u"\u23cf"
        # u"\u23e9"
        # u"\u231a"
        # u"\ufe0f"  # dingbats
        # u"\u3030"
                      # "]+", re.UNICODE)
    # return re.sub(emoj, '', data)
	
def strip_emoji(text):
    # print(emoji.emoji_count(text))
    new_text = re.sub(emoji.get_emoji_regexp(), r"", text)
    return new_text
	
# ------------------------------------------------------
# INPUTS
# ------------------------------------------------------

# Define border
border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

# Define path
path = os.path.abspath(os.path.dirname(__file__))
os.chdir(path)

# ------------------------------------------------------
# Setup worksheet
# ------------------------------------------------------

# Get date
today = str(datetime.datetime.today())[:10]

# Open the workbook to write to
wb = openpyxl.Workbook()

# Open an outfile to write to
fout = open(f'..{os.sep}txt Files{os.sep}{today}.txt', 'w', encoding='utf-8')

# Create the sheets to work in
wb.create_sheet(index=1, title='Properties')
del wb['Sheet']

# Designate sheet to work in
exhibitSheet = wb['Properties']

# Define exhibit title
exhibitTitle = 'Cheap properties from around the world'
# Merge cells for title.
exhibitSheet.merge_cells('A1:C1') 
# Write title to file
exhibitSheet['A1'] = exhibitTitle
# Set format to wrapped
exhibitSheet.cell(row = 1, column = 1).alignment = Alignment(wrap_text = True, vertical='center')
# Set height
exhibitSheet.row_dimensions[1].height = 29
# Set font
fontStyle(exhibitSheet['A1'])
# Remove border
exhibitSheet.cell(row = 1, column = 1).border = None

# Define the header lines
rowText = ['Listing Title', 'Price', 'URL']

# # Split input into list
# rowText = headerExhibit.split('\t')

# Loop through list to write to file
column = 1
rowNum = 2
for item in rowText:
	exhibitSheet.cell(row=rowNum, column=column).value = item
	exhibitSheet.cell(row=rowNum, column=column).alignment = Alignment(horizontal='center', vertical='center')
	fontStyle(exhibitSheet.cell(row=rowNum, column=column))
	column += 1
		
# Set height for header row
exhibitSheet.row_dimensions[2].height = 46

# Define rowWidths
rowWidths = [40, 10.14, 15]

# Adjust columns widths and wrap header text
colNum = 0
# Loop through headers
for col in rowText:
	# Set column width
	exhibitSheet.column_dimensions[string.ascii_uppercase[colNum]].width = rowWidths[colNum]
	# Set format to wrapped
	exhibitSheet.cell(row = 2, column = colNum+1).alignment = Alignment(wrap_text = True, horizontal='center', vertical='center')
	colNum += 1

rowNum = 3

# ------------------------------------------------------
# Get Data
# ------------------------------------------------------

# url = 'https://boston.craigslist.org/search/sof'
locations = {
		'Bologna': {'Nation': 'Italy', 'url': 'https://bologna.craigslist.org/d/housing-real-estate/search/rea?lang=en&cc=gb'}, 
		'Florence': {'Nation': 'Italy', 'url': 'https://florence.craigslist.org/d/housing-real-estate/search/rea?lang=en&cc=gb'}, 
		'Genoa': {'Nation': 'Italy', 'url': 'https://genoa.craigslist.org/d/housing-real-estate/search/rea?lang=en&cc=gb'}, 
		'Milan': {'Nation': 'Italy', 'url': 'https://milan.craigslist.org/d/housing-real-estate/search/rea?lang=en&cc=gb'}, 
		'Napoli': {'Nation': 'Italy', 'url': 'https://naples.craigslist.org/d/housing-real-estate/search/rea?lang=en&cc=gb'}, 
		'Perugia': {'Nation': 'Italy', 'url': 'https://perugia.craigslist.org/d/housing-real-estate/search/rea?lang=en&cc=gb'}, 
		'Rome': {'Nation': 'Italy', 'url': 'https://rome.craigslist.org/d/housing-real-estate/search/rea?lang=en&cc=gb'}, 
		'Sardinia': {'Nation': 'Italy', 'url': 'https://sardinia.craigslist.org/d/housing-real-estate/search/rea?lang=en&cc=gb'}, 
		'Sicilia': {'Nation': 'Italy', 'url': 'https://sicily.craigslist.org/d/housing-real-estate/search/rea?lang=en&cc=gb'}, 
		'Torino': {'Nation': 'Italy', 'url': 'https://torino.craigslist.org/d/housing-real-estate/search/rea?lang=en&cc=gb'}, 
		'Venice': {'Nation': 'Italy', 'url': 'https://venice.craigslist.org/d/housing-real-estate/search/rea?lang=en&cc=gb'}, 
		'Greece': {'Nation': 'Greece', 'url': 'https://athens.craigslist.org/d/real-estate/search/rea'}, 
		'Lisbon': {'Nation': 'Portugal', 'url': 'https://lisbon.craigslist.org/d/housing-real-estate/search/rea?lang=en&cc=gb'}, 
		'Porto': {'Nation': 'Portugal', 'url': 'https://porto.craigslist.org/d/housing-real-estate/search/rea?lang=en&cc=gb'}, 
		'Valencia': {'Nation': 'Spain', 'url': 'https://valencia.craigslist.org/search/rea?lang=en&cc=us'}, 
		'Sevilla': {'Nation': 'Spain', 'url': 'https://sevilla.craigslist.org/search/rea?lang=en&cc=us'}, 
		'Alicante': {'Nation': 'Spain', 'url': 'https://alicante.craigslist.org/search/rea?lang=en&cc=us'}, 
		'Baleares': {'Nation': 'Spain', 'url': 'https://baleares.craigslist.org/search/rea?lang=en&cc=us'}, 
		'Barcelona': {'Nation': 'Spain', 'url': 'https://barcelona.craigslist.org/search/rea?lang=en&cc=us'}, 
		'Bilbao': {'Nation': 'Spain', 'url': 'https://bilbao.craigslist.org/search/rea?lang=en&cc=us'}, 
		'Cadiz': {'Nation': 'Spain', 'url': 'https://cadiz.craigslist.org/search/rea?lang=en&cc=us'}, 
		'Canarias': {'Nation': 'Spain', 'url': 'https://canarias.craigslist.org/search/rea?lang=en&cc=us'}, 
		'Granada': {'Nation': 'Spain', 'url': 'https://granada.craigslist.org/search/rea?lang=en&cc=us'}, 
		'Madrid': {'Nation': 'Spain', 'url': 'https://madrid.craigslist.org/search/rea?lang=en&cc=us'}, 
		'Malaga': {'Nation': 'Spain', 'url': 'https://malaga.craigslist.org/search/rea?lang=en&cc=us'}, 
		'Puerto Rico': {'Nation': 'Puerto Rico', 'url': 'https://puertorico.craigslist.org/d/inmobiliaria/search/rea'}, 
		'Puerto Penasco': {'Nation': 'Mexico', 'url': 'https://phoenix.craigslist.org/search/rea?query=Rocky+point+Mexico&sort=rel&availabilityMode=0&sale_date=all+dates'}, 
		'Dominican Republic': {'Nation': 'Dominican Republic', 'url': 'https://santodomingo.craigslist.org/d/inmobiliaria/search/rea?'}, 
		'Medellin': {'Nation': 'Colombia', 'url': 'https://colombia.craigslist.org/d/real-estate/search/rea?lang=en&cc=us'}, 
		'Belo Horizonte': {'Nation': 'Brazil', 'url': 'https://belohorizonte.craigslist.org/search/rea?'}, 
		'Curitiba': {'Nation': 'Brazil', 'url': 'https://curitiba.craigslist.org/d/im%C3%B3veis/search/rea?'}, 
		'Forteleza': {'Nation': 'Brazil', 'url': 'https://fortaleza.craigslist.org/d/im%C3%B3veis/search/rea?'},  
		'Rio': {'Nation': 'Brazil', 'url': 'https://rio.craigslist.org/d/im%C3%B3veis/search/rea?'}, 
		'Salvador': {'Nation': 'Brazil', 'url': 'https://salvador.craigslist.org/d/im%C3%B3veis/search/rea?'}, 
		'Manila': {'Nation': 'Philippines', 'url': 'https://manila.craigslist.org/d/real-estate/search/rea?'}, 
		'Cagayan de Oro': {'Nation': 'Philippines', 'url': 'https://cdo.craigslist.org/d/real-estate/search/rea?'}, 
		'Cebu': {'Nation': 'Philippines', 'url': 'https://cebu.craigslist.org/d/real-estate/search/rea?'}, 
		'Davao City': {'Nation': 'Philippines', 'url': 'https://davaocity.craigslist.org/d/real-estate/search/rea?'}, 
		'Pampanga': {'Nation': 'Philippines', 'url': 'https://pampanga.craigslist.org/d/real-estate/search/rea?'}, 
		'Panama': {'Nation': 'Panama', 'url': 'https://panama.craigslist.org/search/rea?lang=en&cc=us'}, 
		'Sao Paulo': {'Nation': 'Brazil', 'url': 'https://saopaulo.craigslist.org/d/im%C3%B3veis/search/rea?'}, 
		'Albuquerque': {'Nation': 'United States', 'url': 'https://albuquerque.craigslist.org/d/real-estate/search/rea?'}, 
		'Pittsburgh': {'Nation': 'United States', 'url': 'https://pittsburgh.craigslist.org/d/real-estate/search/rea?'}, 
		'Tucson': {'Nation': 'United States', 'url': 'https://tucson.craigslist.org/d/real-estate/search/rea?'}, 
		}
	
# # Create CurrencyRates object
# c = CurrencyRates()
# # Build currency conversion dictionary
# currency = c.get_rates('USD')
	
# # Add Dominican pesos- price as of 12/30/2020
# currency['DOM'] = 57.17
# # currency['HF'] = 0.0035
# currency['TND'] = 2.68
# currency['COL'] = 3422

# Define url for currency exchange 
url = 'https://api.exchangerate-api.com/v4/latest/USD'

# Get data from API
data = requests.get(url).json()

# Create currency dictionary
currency = data['rates']

# Add Dominican pesos- price as of 5/27/2021
currency['DOM'] = 57.04
currency['COL'] = 3745

# conversions = {
				# 'DOP': '',
				
				# }
				
# # Populate conversions
# for conversion in conversions:
	# # Get response
	# response = requests.get('https://www.xe.com/currencyconverter/convert/?Amount=1&From=USD&To='+conversion)

	# # Create text from reponse
	# data = response.text

	# # Make soup from data
	# soup = BeautifulSoup(data,'html.parser')

	# # Get listItems
	# conversionTag = soup.find_all('span', {'class': 'converterresult-toAmount'})
	
# Loop through urls
for location in locations:
	
	print(f'Searching {location.capitalize()} for deals')
	
	# Add entry to table
	# Define exhibit title
	exhibitTitle = 'Cheap properties from around the world'
	# Merge cells for title.
	# exhibitSheet.merge_cells('A1:C1') 
	exhibitSheet.merge_cells('A'+str(rowNum)+':C'+str(rowNum)) 
				
	# Write title to file
	exhibitSheet['A'+str(rowNum)] = location.capitalize()
	# Set format to wrapped
	exhibitSheet.cell(row = 1, column = 1).alignment = Alignment(wrap_text = True, vertical='center', horizontal='center')
	# # Set height
	# exhibitSheet.row_dimensions[1].height = 29
	# Set font
	fontStyle(exhibitSheet['A'+str(rowNum)])

	# Increment rowNum
	rowNum += 1
	
	# Get response
	response = requests.get(locations[location]['url'])

	# Create text from reponse
	data = response.text

	# Make soup from data
	soup = BeautifulSoup(data,'html.parser')

	# Check if there are listings 
	if soup.find_all('span', {'class': 'button pagenum'})[0].text == 'no results' or soup.find_all('span', {'class': 'button pagenum'})[0].text == 'sem resultado':
		continue
		
	# Get listItems
	listItems = soup.find_all('li', {'class': 'result-row'})
	# tags = soup.find_all('a', {'class': 'result-title hdrlnk'})

	# for tag in tags:
		# print('Job: '+tag.string)
		# print('URL: '+tag.get('href')+'\n')
		
	# Get number of listings 
	listingTotal = int(soup.find_all('span', {'class': 'rangeTo'})[0].text)
	# Check if listingTotal is 120
	if listingTotal == 120:
		# Get totalcount
		listingTotal = int(soup.find_all('span', {'class': 'totalcount'})[0].text)
		
	listingNum = 0
	
	# Check if there are more listings than displayed on the page
	if listingTotal > listingNum:
		print('There are more listing than fit on one page. Adding another page')
		
		# Generate link to next page
		nextPage = f"{locations[location]['url'].replace('rea?', 'rea?s=120&')}"
		
		# * Do something with next page
		# ** -> Functionalize things so that the next page can be examined too 
		
	# Loop through listItems
	for listItem in listItems:
		
		# Increment listingNum
		listingNum += 1
		print(f'Compiling {listingNum} of {listingTotal} listings')
		
		# Check if we have viewed as many properties are are available locally 
		if listingNum > listingTotal:
			break 
			
		# print(listItem)
		price = listItem.find_all('span', {'class': 'result-price'})[0].text
		# print(price)
		orPrice = price
		
		# Check if price is dollars
		if price[0] == u"$":
			price = float(price[1:].replace(",",""))
			conversion = 1
			
		# Check if price is isn't listed
		elif len(price) <= 2:
			continue
			
		# Check if price is euros
		elif price[0] == u"\N{euro sign}":
		
			# Get conversion
			conversion = currency['EUR']
			
			# Convert price from euros string to euros float
			price = float(price[1:].replace(".", "").replace('\xa0', '').replace(",",""))
			
		elif price[:3] == u"CHF":
		
			# Get conversion
			conversion = currency['CHF']
			
			# Convert price from euros string to euros float
			price = float(price[3:].replace(".", "").replace('\xa0', '').replace(",","."))
		
		elif price[:3] == u"TND":
		
			# Get conversion
			conversion = currency['TND']
			
			# Convert price from euros string to euros float
			price = float(price[3:].replace(".", "").replace('\xa0', '').replace(",","."))
		
		elif price[:3] == u"HRK":
		
			# Get conversion
			conversion = currency['HRK']
			
			# Convert price from euros string to euros float
			price = float(price[3:].replace(".", "").replace('\xa0', '').replace(",",""))
		
		elif price[:3] == u"COL":
		
			# Get conversion
			conversion = currency['COL']
			
			# Convert price from euros string to euros float
			price = float(price[3:].replace(".", "").replace('\xa0', '').replace(",",""))
		
		elif price[:2] == u"R$":
		
			# Get conversion
			conversion = currency['BRL']
			
			# Convert price from euros string to euros float
			price = float(price[2:].replace(".", "").replace('\xa0', '').replace(",",""))
		
		elif location == 'Dominican Republic' or location == 'Puerto Rico':
		
			# print(price)
			# Check if price is $
			if price[0] != '$':
				# Get conversion
				conversion = currency['DOM']
			
			else:
				conversion = 1
				
			# Convert price from euros string to euros float
			price = float(price[1:].replace(".", "").replace('\xa0', '').replace(",",""))
			# print(price)
		
		elif location == 'Medellin':
			# print(price)
			# Get conversion
			conversion = currency['COL']
		
		elif locations[location]['Nation'] == 'Philippines':
			# print(price)
			# Convert price from euros string to euros float
			price = float(price[1:].replace(".", "").replace('\xa0', '').replace(",",""))
			
			# Get conversion
			conversion = currency['PHP']
		
		# elif price[0] not in [u"\N{euro sign}", "$", 'C', 'H', 'T'] 'CHF', 'TND', 'HRK', 'COL', 'R$']:
		# # if url == 'Dominican Republic':
			# print(f'Price is in Dominican pesos {price}')
		
			# # Get conversion
			# conversion = currency['DOM']
		
			# # Convert price from euros string to euros float
			# price = float(price[1:].replace(".", "").replace('\xa0', '').replace(",","."))

		# -------------------------------------------------------
		
		# Convert to dollars
		price = price/conversion
		# print(f'Price converted to dollars {price}')
		
		# Get all links
		links = listItem.find_all(href=True)
		# Get link
		link = links[0]['href']
		# Get title
		title = str(links[1].find_all(text=True)[0])

		# Remove emojis
		title = strip_emoji(title)
		# Remove /
		title = title.replace('/','-').replace('\t','-')
		# bytes(title, 'utf-8').decode('utf-8','ignore')
		# bytes(weird, 'utf-8').decode('utf-8','ignore')
		if '𝐓𝐢𝐫𝐞𝐝 𝐎𝐟 𝐏𝐚𝐲𝐢𝐧𝐠 𝐑𝐞𝐧𝐭' in title:
			weird = title
			print(title)
			continue
			
		# Write out file for DB import
		fout.write(f"{location}\t{locations[location]['Nation']}\t{title}\t{round(price)}\t{link}\t{today}\n")
		
		# Check if less than threshold
		if 1500 < price and price <= 50000:
			
		
			# Package info into list
			info = [title, round(price), link]
			# print(f'{title}\n{price}\n{link}\n')
			
			column = 1
			# Loop through headers
			for item in info:
				exhibitSheet.cell(row=rowNum, column=column).value = item
				# Make urls hyperlinked
				if column == 3:
					fontStyleURL(exhibitSheet.cell(row=rowNum, column=column), link)
					# exhibitSheet.cell(row=rowNum, column=column).style = "Hyperlink"
				else:
					fontStyle(exhibitSheet.cell(row=rowNum, column=column))
				exhibitSheet.cell(row=rowNum, column=column).alignment = Alignment(wrap_text = True, horizontal='center', vertical='center')
				column += 1
			rowNum += 1

	# print(f'{listingTotal} properties found')
	# if listingTotal > 119:
		# print(dog)
	
# Close out file
fout.close()

# Write to excel file
wb.save(f'..{os.sep}Excel Files{os.sep}Discount Properties- {today}.xlsx')

# Excel formula to build urls
# ="'"&A10&"': '"&B10&"', "